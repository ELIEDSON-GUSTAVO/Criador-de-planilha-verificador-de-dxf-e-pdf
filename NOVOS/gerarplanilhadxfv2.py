import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import win32com.client as win32  # Importando a biblioteca para controlar o Outlook

def processar_arquivo():
    # Seleciona o arquivo de entrada com um clique
    caminho_entrada = filedialog.askopenfilename(
        title="Selecione o arquivo Excel de entrada",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    if not caminho_entrada:
        return

    try:
        # Lê a aba correta da planilha
        df = pd.read_excel(caminho_entrada, sheet_name='Lista de materiais')

        # Seleciona colunas relevantes
        colunas = ['Número da peça', 'QTDE', 'Descrição', 'Massa', 'Material', 'Caminho do arquivo']
        df_filtrado = df[colunas].copy()

        # Corrige os caminhos do arquivo
        df_filtrado['Caminho do arquivo'] = df_filtrado['Caminho do arquivo'].str.replace(
            'T:\\14 - PDF\\Engenharia', 'T:\\17 - DXF\\Engenharia', regex=False)
        df_filtrado['Caminho do arquivo'] = df_filtrado['Caminho do arquivo'].str.replace(
            '.idw.pdf', '.ipt.dxf', regex=False)

        # Aplica os filtros
        cond_topo = df_filtrado['Número da peça'].astype(str).str.startswith(('02.01.01', '02.01.02'))
        cond_exclusao = df_filtrado['Material'].astype(str).str.startswith((
            '01.04.03', '01.04.04', '01.05.06', '01.04.06', '01.02.01', '01.05.04'))

        df_topo = df_filtrado[cond_topo & ~cond_exclusao].copy()
        df_outros = df_filtrado[~(cond_topo & ~cond_exclusao)].copy()

        # Marca os que devem ficar em amarelo (opcional, para visualização)
        df_topo['destacar_amarelo'] = False
        df_outros['destacar_amarelo'] = True

        # Junta os dois grupos
        df_final = pd.concat([df_topo, df_outros], ignore_index=True)

        # Seleciona onde salvar o novo arquivo
        caminho_saida = filedialog.asksaveasfilename(
            title="Salvar como...",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if not caminho_saida:
            return

        # Salva o novo Excel com destaque nas células
        df_final.to_excel(caminho_saida, index=False)

        # Carrega o arquivo para formatação
        wb = load_workbook(caminho_saida)
        sheet = wb.active

        # Aplica o destaque amarelo nas células
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in range(2, sheet.max_row + 1):  # Começa na linha 2 para não formatar o cabeçalho
            if sheet.cell(row=row, column=df_final.columns.get_loc('destacar_amarelo') + 1).value:
                for col in range(1, len(df_final.columns) + 1):
                    sheet.cell(row=row, column=col).fill = fill

        # Salva as modificações no arquivo Excel
        wb.save(caminho_saida)

        messagebox.showinfo("Sucesso", "Arquivo processado e salvo com sucesso!")

        # Após salvar o arquivo, envia o e-mail
        EnviarComAssinaturaOutlook(caminho_saida)

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro:\n{str(e)}")


def EnviarComAssinaturaOutlook(caminho_arquivo):
    try:
        OutlookApp = win32.Dispatch("Outlook.Application")
        OutlookMail = OutlookApp.CreateItem(0)

        # Texto do e-mail (antes da assinatura)
        CorpoSemAssinatura = "<p>Olá,</p>" + \
                             "<p>Segue em anexo a planilha do BOM - Projeto ALMA.</p><br>"

        # Preenche as propriedades do e-mail
        OutlookMail.To = "cadastro.produtos@nextimplementos.com.br"
        OutlookMail.Subject = "Envio de Planilha - BOM ALMA"
        OutlookMail.Display()  # Exibe o e-mail antes de enviar (necessário para carregar a assinatura)

        # Junta o corpo do e-mail com a assinatura padrão do Outlook
        OutlookMail.HTMLBody = CorpoSemAssinatura + OutlookMail.HTMLBody

        # Anexa a planilha salva
        OutlookMail.Attachments.Add(caminho_arquivo)

        # Não envia o e-mail, apenas abre para revisão
        # OutlookMail.Send()  # Esta linha foi comentada
        messagebox.showinfo("Sucesso", "E-mail preparado. Revise e envie manualmente!")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao preparar o e-mail:\n{str(e)}")


# Interface com Tkinter
janela = tk.Tk()
janela.title("Filtrador de Lista de Materiais")
janela.geometry("420x160")

btn = tk.Button(janela, text="Selecionar e Processar Arquivo Excel",
                command=processar_arquivo, font=("Arial", 12), padx=10, pady=10)
btn.pack(pady=40)

janela.mainloop()
